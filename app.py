"""
Mutual Fund Portfolio Dashboard — v2
Premium-grade analytics: FIFO capital gains, tax preview, allocation drift, SIP health.
Inspired by Groww, mProfit, ET Money, MF Central, Paytm Money.
"""
from __future__ import annotations

import io
import json
from collections import defaultdict
from dataclasses import asdict
from datetime import date, datetime
from typing import Optional

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

from cas_parser import (
    ParsedCAS, Scheme, Transaction, parse_cas, xirr, scheme_external_flows,
)
from analytics import (
    TaxRates, SchemeCategory, GainsLot, UnrealizedLot, TaxSummary, DriftRow,
    SipInfo,
    categorize_scheme, compute_capital_gains, compute_tax_for_lots,
    simulate_full_redemption_tax, compute_drift, analyze_sips,
    canonical_holder_per_pan,
)
from database import init_db, get_db, save_portfolio, list_portfolios, load_portfolio, delete_portfolio, portfolio_to_session_state


# ==================== Page setup ====================
if "db_initialized" not in st.session_state:
    try:
        init_db()
    except Exception:
        pass
st.session_state.db_initialized = True
st.set_page_config(
    page_title="MF Portfolio Dashboard",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
    .stApp { background: #f8fafc; }
    .main .block-container { padding-top: 1rem; padding-bottom: 2rem; max-width: 1500px; }
    h1 { color: #0f172a; font-weight: 700; font-size: 2rem; margin-bottom: 0.25rem; }
    h2 { color: #1e293b; font-weight: 600; margin-top: 1.5rem; }
    h3 { color: #334155; font-weight: 600; }
    h4 { color: #475569; font-weight: 600; margin-top: 1rem; }
    .stTabs [data-baseweb="tab-list"] { gap: 4px; flex-wrap: wrap; }
    .stTabs [data-baseweb="tab"] {
        background: white; border: 1px solid #e2e8f0; border-radius: 8px;
        padding: 8px 14px; font-size: 0.92rem;
    }
    .stTabs [aria-selected="true"] {
        background: #2563eb !important; color: white !important; border-color: #2563eb;
    }
    div[data-testid="stMetric"] {
        background: white; padding: 1rem; border-radius: 12px;
        box-shadow: 0 1px 3px rgba(0,0,0,0.05); border: 1px solid #e2e8f0;
    }
    div[data-testid="stMetric"] label { color: #64748b; font-size: 0.78rem; }
    div[data-testid="stMetricValue"] { color: #0f172a; font-weight: 700; }
    .insight {
        background: #eff6ff; border-left: 4px solid #2563eb;
        padding: 0.75rem 1rem; border-radius: 6px; margin: 0.5rem 0;
        color: #1e40af; font-size: 0.92rem;
    }
    .warning {
        background: #fef3c7; border-left: 4px solid #f59e0b;
        padding: 0.75rem 1rem; border-radius: 6px; margin: 0.5rem 0;
        color: #92400e; font-size: 0.92rem;
    }
</style>
""", unsafe_allow_html=True)


# ==================== Session state ====================
def init_state():
    defaults = {
        "parsed_files": [],          # list of (filename, ParsedCAS)
        "pan_holders": {},
        "tax_rates": TaxRates(),
        "debt_slab": 30.0,
        "allocation_targets": {"Equity": 70.0, "Debt": 20.0, "Hybrid": 5.0, "Gold": 5.0},
        "scheme_overrides": {},      # scheme_name -> {"asset_class": ..., "tax_category": ...}
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

init_state()


# ==================== Helpers ====================
@st.cache_data(show_spinner=False)
def parse_pdf_cached(file_bytes: bytes, file_name: str, password: Optional[str]):
    cas = parse_cas(file_bytes, password=password or None, source_filename=file_name)
    return cas


def fmt_inr(x, decimals=0):
    if x is None or pd.isna(x):
        return "—"
    sign = "-" if x < 0 else ""
    x = abs(x)
    if x >= 1e7:
        return f"{sign}₹{x/1e7:,.2f}Cr"
    if x >= 1e5:
        return f"{sign}₹{x/1e5:,.2f}L"
    if decimals == 0:
        return f"{sign}₹{x:,.0f}"
    return f"{sign}₹{x:,.{decimals}f}"


def fmt_pct(x):
    if x is None or pd.isna(x):
        return "—"
    return f"{x:+.2f}%"


def all_schemes() -> list:
    """Combine schemes from all uploaded CAS files, deduped by (PAN, folio, scheme, AMC)."""
    seen = {}
    for _, cas in st.session_state.parsed_files:
        for s in cas.schemes:
            key = (s.pan, s.folio, s.scheme_name, s.amc)
            existing = seen.get(key)
            if existing is None or len(s.cashflows) > len(existing.cashflows):
                seen[key] = s
    return list(seen.values())


def schemes_to_df(schemes: list, holders: dict) -> pd.DataFrame:
    rows = []
    for s in schemes:
        cat = get_category(s)
        flows = scheme_external_flows(s)
        r = xirr(flows)
        rows.append({
            "PAN": s.pan,
            "Holder": holders.get(s.pan, s.holder),
            "AMC": s.amc,
            "Scheme": s.scheme_name,
            "Folio": s.folio,
            "Asset_Class": cat.asset_class,
            "Sub_Category": cat.sub_category,
            "Tax_Category": cat.tax_category,
            "Cost": s.cost,
            "Market_Value": s.market_value,
            "Gain": s.market_value - s.cost,
            "Abs_Return_Pct": (s.market_value / s.cost - 1) * 100 if s.cost else 0,
            "XIRR_Pct": (r * 100) if r is not None else None,
            "NAV_Date": s.nav_date,
            "Closing_Units": s.closing_units,
            "Transactions": len(s.cashflows),
            "First_Investment": min((t.date for t in s.cashflows), default=None),
            "SIP_Status": s.sip_status,
            "Source_File": s.source_file,
        })
    return pd.DataFrame(rows)


def get_category(s: Scheme) -> SchemeCategory:
    """Get category for a scheme, applying user overrides."""
    base = categorize_scheme(s.scheme_name)
    override = st.session_state.scheme_overrides.get(s.scheme_name)
    if override:
        return SchemeCategory(
            asset_class=override.get("asset_class", base.asset_class),
            sub_category=override.get("sub_category", base.sub_category),
            tax_category=override.get("tax_category", base.tax_category),
        )
    return base


def aggregate_xirr(schemes: list, group_fn) -> dict:
    groups = defaultdict(list)
    for s in schemes:
        gk = group_fn(s)
        flows = scheme_external_flows(s)
        groups[gk].extend(flows)
    return {gk: (xirr(fl) * 100 if xirr(fl) is not None else None) for gk, fl in groups.items()}


def export_state_json() -> bytes:
    """Serialize session state to JSON for re-import without re-parsing PDFs."""
    out = {"version": 2, "exported_at": datetime.now().isoformat(), "files": []}
    for fname, cas in st.session_state.parsed_files:
        out["files"].append({
            "filename": fname,
            "email": cas.email,
            "statement_start": cas.statement_start.isoformat() if cas.statement_start else None,
            "statement_end": cas.statement_end.isoformat() if cas.statement_end else None,
            "source_filename": cas.source_filename,
            "schemes": [{
                "pan": s.pan, "holder": s.holder, "amc": s.amc,
                "folio": s.folio, "scheme_name": s.scheme_name,
                "cost": s.cost, "market_value": s.market_value,
                "nav_date": s.nav_date.isoformat() if s.nav_date else None,
                "closing_units": s.closing_units,
                "sip_status": s.sip_status,
                "sip_cancel_date": s.sip_cancel_date.isoformat() if s.sip_cancel_date else None,
                "source_file": s.source_file,
                "cashflows": [{
                    "date": t.date.isoformat(), "kind": t.kind,
                    "amount": t.amount, "units": t.units, "price": t.price,
                    "description": t.description,
                } for t in s.cashflows],
            } for s in cas.schemes],
        })
    out["targets"] = st.session_state.allocation_targets
    out["debt_slab"] = st.session_state.debt_slab
    return json.dumps(out, indent=2).encode("utf-8")


def import_state_json(data: bytes):
    """Restore session state from previously-exported JSON."""
    obj = json.loads(data.decode("utf-8"))
    files = []
    for f in obj.get("files", []):
        cas = ParsedCAS(
            email=f.get("email", ""),
            statement_start=date.fromisoformat(f["statement_start"]) if f.get("statement_start") else None,
            statement_end=date.fromisoformat(f["statement_end"]) if f.get("statement_end") else None,
            source_filename=f.get("source_filename", ""),
        )
        for sd in f.get("schemes", []):
            sch = Scheme(
                pan=sd["pan"], holder=sd["holder"], amc=sd["amc"],
                folio=sd["folio"], scheme_name=sd["scheme_name"],
                cost=sd["cost"], market_value=sd["market_value"],
                nav_date=date.fromisoformat(sd["nav_date"]) if sd.get("nav_date") else None,
                closing_units=sd.get("closing_units", 0),
                sip_status=sd.get("sip_status", ""),
                sip_cancel_date=date.fromisoformat(sd["sip_cancel_date"]) if sd.get("sip_cancel_date") else None,
                source_file=sd.get("source_file", ""),
            )
            for t in sd.get("cashflows", []):
                sch.cashflows.append(Transaction(
                    date=date.fromisoformat(t["date"]), kind=t["kind"],
                    amount=t["amount"], units=t.get("units", 0),
                    price=t.get("price", 0), description=t.get("description", ""),
                ))
            cas.schemes.append(sch)
        files.append((f["filename"], cas))
    st.session_state.parsed_files = files
    if "targets" in obj:
        st.session_state.allocation_targets = obj["targets"]
    if "debt_slab" in obj:
        st.session_state.debt_slab = obj["debt_slab"]
    schemes = all_schemes()
    st.session_state.pan_holders = canonical_holder_per_pan(schemes)


def export_excel() -> bytes:
    """Create a multi-sheet Excel workbook with all key tables."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    schemes = all_schemes()
    holders = st.session_state.pan_holders or canonical_holder_per_pan(schemes)
    df = schemes_to_df(schemes, holders)

    wb = Workbook()

    HEADER_FILL = PatternFill("solid", start_color="1E40AF")
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    THIN = Side(border_style="thin", color="BFBFBF")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def write_sheet(ws, df_out, money_cols=None, pct_cols=None):
        money_cols = money_cols or set()
        pct_cols = pct_cols or set()
        cols = list(df_out.columns)
        for j, col in enumerate(cols, 1):
            c = ws.cell(row=1, column=j, value=col.replace("_", " "))
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = BORDER
        for i, row in enumerate(df_out.itertuples(index=False), 2):
            for j, (col, val) in enumerate(zip(cols, row), 1):
                c = ws.cell(row=i, column=j, value=val)
                c.font = Font(name="Arial", size=10)
                c.border = BORDER
                if col in money_cols:
                    c.number_format = '#,##0;(#,##0);"-"'
                elif col in pct_cols:
                    c.number_format = '0.00"%";(0.00"%");"-"'
        for j, col in enumerate(cols, 1):
            try:
                width = max([len(str(col))] + [len(str(v)[:30]) for v in df_out[col].head(50)])
                ws.column_dimensions[get_column_letter(j)].width = min(width + 3, 50)
            except Exception:
                ws.column_dimensions[get_column_letter(j)].width = 15
        ws.freeze_panes = "A2"

    # Sheet 1: Per-PAN summary
    ws = wb.active
    ws.title = "By PAN"
    pan = df.groupby("PAN").agg(
        Holder=("Holder", "first"), Schemes=("Scheme", "count"),
        Cost=("Cost", "sum"), Market_Value=("Market_Value", "sum"),
    ).reset_index()
    pan["Gain"] = pan["Market_Value"] - pan["Cost"]
    pan["Abs_Return_Pct"] = (pan["Market_Value"] / pan["Cost"] - 1) * 100
    pan_xirr = aggregate_xirr(schemes, lambda s: s.pan)
    pan["XIRR_Pct"] = pan["PAN"].map(pan_xirr)
    write_sheet(ws, pan, {"Cost", "Market_Value", "Gain"}, {"Abs_Return_Pct", "XIRR_Pct"})

    # Sheet 2: PAN x AMC
    ws = wb.create_sheet("By PAN x AMC")
    pa = df.groupby(["PAN", "AMC"]).agg(
        Holder=("Holder", "first"), Schemes=("Scheme", "count"),
        Cost=("Cost", "sum"), Market_Value=("Market_Value", "sum"),
    ).reset_index()
    pa["Gain"] = pa["Market_Value"] - pa["Cost"]
    pa["Abs_Return_Pct"] = (pa["Market_Value"] / pa["Cost"] - 1) * 100
    pa_xirr = aggregate_xirr(schemes, lambda s: (s.pan, s.amc))
    pa["XIRR_Pct"] = pa.apply(lambda r: pa_xirr.get((r["PAN"], r["AMC"])), axis=1)
    write_sheet(ws, pa, {"Cost", "Market_Value", "Gain"}, {"Abs_Return_Pct", "XIRR_Pct"})

    # Sheet 3: Per Scheme
    ws = wb.create_sheet("Per Scheme")
    write_sheet(ws, df, {"Cost", "Market_Value", "Gain"}, {"Abs_Return_Pct", "XIRR_Pct"})

    # Sheet 4: Asset Allocation
    ws = wb.create_sheet("Asset Allocation")
    aa = df.groupby("Asset_Class").agg(
        Schemes=("Scheme", "count"), Cost=("Cost", "sum"),
        Market_Value=("Market_Value", "sum"),
    ).reset_index()
    aa["Allocation_Pct"] = aa["Market_Value"] / aa["Market_Value"].sum() * 100
    write_sheet(ws, aa, {"Cost", "Market_Value"}, {"Allocation_Pct"})

    # Sheet 5: Realized Gains
    all_realized, all_unrealized = [], []
    for s in schemes:
        r, u = compute_capital_gains(s, st.session_state.tax_rates)
        all_realized.extend(r)
        all_unrealized.extend(u)
    if all_realized:
        ws = wb.create_sheet("Realized Gains")
        rg = pd.DataFrame([{
            "FY": l.fy, "Scheme": l.scheme, "PAN": l.pan, "Holder": l.holder,
            "Buy_Date": l.buy_date, "Sell_Date": l.sell_date,
            "Units": l.units, "Buy_Price": l.buy_price, "Sell_Price": l.sell_price,
            "Cost": l.cost, "Proceeds": l.proceeds, "Gain": l.gain,
            "Holding_Days": l.holding_days, "Term": "LTCG" if l.is_long_term else "STCG",
            "Tax_Category": l.tax_category,
        } for l in all_realized])
        write_sheet(ws, rg, {"Cost", "Proceeds", "Gain"})

    # Sheet 6: Unrealized Gains (summary by scheme)
    if all_unrealized:
        ws = wb.create_sheet("Unrealized Gains")
        ur_by_scheme = defaultdict(lambda: {"cost": 0, "value": 0, "gain": 0,
                                            "ltcg_gain": 0, "stcg_gain": 0,
                                            "pan": "", "holder": "", "amc": "", "tax_cat": ""})
        for l in all_unrealized:
            d = ur_by_scheme[l.scheme]
            d["cost"] += l.cost; d["value"] += l.current_value; d["gain"] += l.gain
            if l.is_long_term: d["ltcg_gain"] += l.gain
            else: d["stcg_gain"] += l.gain
            d["pan"] = l.pan; d["holder"] = l.holder; d["amc"] = l.amc; d["tax_cat"] = l.tax_category
        ur_df = pd.DataFrame([{
            "Scheme": k, "PAN": v["pan"], "Holder": v["holder"], "AMC": v["amc"],
            "Tax_Category": v["tax_cat"], "Cost": v["cost"], "Current_Value": v["value"],
            "Total_Gain": v["gain"], "LTCG_Gain": v["ltcg_gain"], "STCG_Gain": v["stcg_gain"],
        } for k, v in ur_by_scheme.items()])
        write_sheet(ws, ur_df, {"Cost", "Current_Value", "Total_Gain", "LTCG_Gain", "STCG_Gain"})

    # Sheet 7: Tax by FY
    if all_realized:
        ws = wb.create_sheet("Tax by FY")
        tax_summary = compute_tax_for_lots(all_realized, st.session_state.tax_rates,
                                          st.session_state.debt_slab / 100)
        tax_df = pd.DataFrame([{
            "FY": fy, "Total_Gain": s.total_gain,
            "Equity_STCG": s.equity_stcg, "Equity_LTCG": s.equity_ltcg_gross,
            "Equity_LTCG_Taxable": s.equity_ltcg_taxable,
            "Debt_STCG": s.debt_stcg, "Debt_LTCG": s.debt_ltcg,
            "Tax_Equity_STCG": s.equity_stcg_tax, "Tax_Equity_LTCG": s.equity_ltcg_tax,
            "Tax_Debt_STCG": s.debt_stcg_tax, "Tax_Debt_LTCG": s.debt_ltcg_tax,
            "Total_Tax": s.total_tax,
        } for fy, s in sorted(tax_summary.items())])
        write_sheet(ws, tax_df,
                    {"Total_Gain", "Equity_STCG", "Equity_LTCG", "Equity_LTCG_Taxable",
                     "Debt_STCG", "Debt_LTCG", "Tax_Equity_STCG", "Tax_Equity_LTCG",
                     "Tax_Debt_STCG", "Tax_Debt_LTCG", "Total_Tax"})

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


# ==================== Sidebar ====================
with st.sidebar:
    st.markdown("## 📊 Portfolio Dashboard")
    st.caption("v2 · Multi-PAN · Capital Gains · Tax Preview")
    st.divider()

    st.markdown("### 📁 Import")
    uploaded = st.file_uploader("CAS PDFs", type=["pdf"], accept_multiple_files=True,
                                 help="Upload one or more CAMS+KFintech CAS PDFs")
    pwd = st.text_input("Password (if encrypted)", type="password",
                        placeholder="Usually your PAN")

    cc1, cc2 = st.columns(2)
    if cc1.button("Parse", type="primary", width="stretch", disabled=not uploaded):
        new_list = []
        for f in uploaded:
            try:
                cas = parse_pdf_cached(f.read(), f.name, pwd)
                if cas.schemes:
                    new_list.append((f.name, cas))
                    st.success(f"✓ {f.name[:30]} ({len(cas.schemes)} schemes)")
                else:
                    if cas.parse_warnings:
                        msg = "\n".join(cas.parse_warnings)
                        st.error(f"✗ {f.name[:30]}:\n{msg}")
                    else:
                        st.warning(f"⚠ {f.name[:30]}: No schemes found — is this a valid CAS PDF?")
            except Exception as e:
                st.error(f"✗ {f.name[:30]}: {e}")
        if new_list:
            st.session_state.parsed_files = new_list
            schemes = all_schemes()
            st.session_state.pan_holders = canonical_holder_per_pan(schemes)
            st.rerun()

    if cc2.button("Clear", width="stretch", disabled=not uploaded):
        st.session_state.parsed_files = []
        st.session_state.pan_holders = {}
        st.rerun()

    # JSON state import
    st.markdown("##### Or load saved state")
    json_file = st.file_uploader("JSON state", type=["json"], key="json_uploader",
                                  label_visibility="collapsed")
    if json_file and st.button("Load JSON", width="stretch"):
        try:
            import_state_json(json_file.read())
            st.success("Loaded!")
            st.rerun()
        except Exception as e:
            st.error(f"Load failed: {e}")

    with st.expander("ℹ️ PDF parsing tips"):
        st.markdown("""
        - **Encrypted PDFs**: Use your PAN as password (uppercase)
        - **Wrong CAS type**: Download "Detailed" CAS (PDF), not "Summary"
        - **CAMS+KFintech**: The detailed CAS covers both; single statement is sufficient
        - **Scanned PDFs**: Not supported — request text-based CAS from the registrar
        """)
        if st.button("🗑️ Clear PDF cache", width="stretch"):
            st.cache_data.clear()
            st.success("Cache cleared. Try parsing again.")
        if st.button("↻ Reload PDFs", width="stretch", disabled=not uploaded):
            new_list = []
            for f in uploaded:
                try:
                    cas = parse_pdf_cached(f.read(), f.name, pwd)
                    if cas.schemes:
                        new_list.append((f.name, cas))
                        st.success(f"✓ {f.name[:30]} ({len(cas.schemes)} schemes)")
                    else:
                        if cas.parse_warnings:
                            msg = "\n".join(cas.parse_warnings)
                            st.error(f"✗ {f.name[:30]}:\n{msg}")
                        else:
                            st.warning(f"⚠ {f.name[:30]}: No schemes found — is this a valid CAS PDF?")
                except Exception as e:
                    st.error(f"✗ {f.name[:30]}: {e}")
            if new_list:
                st.session_state.parsed_files = new_list
                schemes = all_schemes()
                st.session_state.pan_holders = canonical_holder_per_pan(schemes)
                st.rerun()

    if st.session_state.parsed_files:
        st.divider()
        st.markdown("### 📂 Loaded files")
        for fname, cas in st.session_state.parsed_files:
            st.caption(f"📄 **{fname[:35]}**")
            st.caption(f"  {cas.email} · {len(cas.schemes)} schemes")

    st.divider()
    st.markdown("### 💾 Saved Portfolios")
    try:
        session = get_db()
        portfolios = list_portfolios(session)
        session.close()
        if portfolios:
            for p in portfolios:
                cols = st.columns([3, 1, 1])
                cols[0].caption(f"📁 {p.name} ({len(p.schemes)} schemes)")
                if cols[1].button("Load", key=f"load_{p.id}"):
                    session = get_db()
                    full = load_portfolio(session, p.id)
                    if full:
                        state = portfolio_to_session_state(full)
                        st.session_state.parsed_files = state["parsed_files"]
                        st.session_state.pan_holders = state["pan_holders"]
                        if state["tax_rates"]:
                            tr = st.session_state.tax_rates
                            tr.equity_ltcg_rate = state["tax_rates"].get("equity_ltcg_rate", tr.equity_ltcg_rate)
                            tr.equity_ltcg_exemption = state["tax_rates"].get("equity_ltcg_exemption", tr.equity_ltcg_exemption)
                            tr.equity_stcg_rate = state["tax_rates"].get("equity_stcg_rate", tr.equity_stcg_rate)
                            tr.debt_ltcg_rate = state["tax_rates"].get("debt_ltcg_rate", tr.debt_ltcg_rate)
                            tr.debt_stcg_slab = state["tax_rates"].get("debt_stcg_slab", tr.debt_stcg_slab)
                        st.session_state.allocation_targets = state["allocation_targets"] or st.session_state.allocation_targets
                        st.session_state.scheme_overrides = state["scheme_overrides"] or st.session_state.scheme_overrides
                        st.rerun()
                if cols[2].button("Del", key=f"del_{p.id}"):
                    session = get_db()
                    delete_portfolio(session, p.id)
                    st.rerun()
        else:
            st.caption("No saved portfolios yet.")
    except Exception as e:
        st.caption(f"DB not available: {e}")


# ==================== Main: empty state ====================
st.title("📊 Mutual Fund Portfolio Dashboard")

if not st.session_state.parsed_files:
    st.caption("Premium-grade analytics for your full family MF portfolio.")
    st.info("👈 Upload one or more CAS PDFs in the sidebar to begin, or load a saved JSON state.")
    with st.expander("ℹ️ What's new in v2", expanded=True):
        st.markdown("""
**Inspired by Groww, mProfit, ET Money Genius, MF Central, Paytm Money** — built free and self-hostable.

**Free-tier parity** (same as those apps' default views):
- Family-wide combined view across multiple PANs and emails
- Asset-class allocation, sector and category breakdowns
- AMC and scheme-level tables with cost, MV, gain, XIRR
- Top/bottom performers, monthly investment trend

**Premium-feature parity** (usually paid in those apps):
- 💼 **Capital Gains** — FIFO-matched realized gains by FY, ITR-ready (mProfit Pro)
- 🎯 **Tax Preview** — what-if redemption tax in real time (Groww Tax-asana)
- ⚖️ **Allocation Drift** — vs your target allocation, with rebalance amounts (ET Money Genius)
- 🔄 **SIP Health** — active / cancelled SIPs, days since last instalment
- 📤 **Excel + JSON export** — full state, re-importable so you don't re-parse PDFs each time

Built for India: rupee formatting (Lakhs/Crores), FY 2024-25 tax regime by default, post-July 2024 LTCG rules.
""")
    st.stop()


# ==================== Build dataframes ====================
schemes = all_schemes()
holders = st.session_state.pan_holders or canonical_holder_per_pan(schemes)
df_master = schemes_to_df(schemes, holders)

# Sidebar filters
with st.sidebar:
    st.divider()
    st.markdown("### 🎚️ Filters")
    pan_options = sorted(df_master["PAN"].unique())
    selected_pans = st.multiselect(
        "PAN", options=pan_options, default=pan_options,
        format_func=lambda p: f"{p} — {holders.get(p, '?')[:25]}",
    )
    amc_options = sorted(df_master["AMC"].unique())
    selected_amcs = st.multiselect("AMC", options=amc_options, default=amc_options)
    asset_options = sorted(df_master["Asset_Class"].unique())
    selected_assets = st.multiselect("Asset Class", options=asset_options, default=asset_options)

df = df_master[
    df_master["PAN"].isin(selected_pans)
    & df_master["AMC"].isin(selected_amcs)
    & df_master["Asset_Class"].isin(selected_assets)
]
filtered_schemes = [
    s for s in schemes
    if s.pan in selected_pans
    and s.amc in selected_amcs
    and get_category(s).asset_class in selected_assets
]

if df.empty:
    st.warning("No schemes match the current filters.")
    st.stop()


# ==================== Top KPIs ====================
total_cost = df["Cost"].sum()
total_mv = df["Market_Value"].sum()
total_gain = total_mv - total_cost
abs_return = (total_mv / total_cost - 1) * 100 if total_cost else 0

all_flows = []
for s in filtered_schemes:
    all_flows.extend(scheme_external_flows(s))
portfolio_xirr = xirr(all_flows)

# Hypothetical tax preview for KPIs
_, hyp_tax, _, _ = simulate_full_redemption_tax(filtered_schemes, st.session_state.tax_rates,
                                                 st.session_state.debt_slab / 100)

c1, c2, c3, c4, c5, c6 = st.columns(6)
c1.metric("Invested", fmt_inr(total_cost))
c2.metric("Current Value", fmt_inr(total_mv))
c3.metric("Gain", fmt_inr(total_gain), delta=f"{abs_return:+.2f}%", delta_color="normal")
c4.metric("Portfolio XIRR", f"{portfolio_xirr*100:.2f}%" if portfolio_xirr else "—")
c5.metric("Tax if redeem now", fmt_inr(hyp_tax),
          delta=f"-{hyp_tax/total_mv*100:.2f}%" if total_mv else "—",
          delta_color="inverse")
c6.metric("Holdings", f"{len(df)}", delta=f"{df['PAN'].nunique()} PAN · {df['AMC'].nunique()} AMC",
          delta_color="off")


# ==================== Tabs ====================
tabs = st.tabs([
    "🏠 Overview", "👥 Family", "🏢 By AMC", "📋 Schemes",
    "💼 Capital Gains", "🎯 Tax Preview", "⚖️ Allocation",
    "🔄 SIP Health", "💸 Cash Flows", "📤 Export",
])


# -------------------- 1. OVERVIEW --------------------
with tabs[0]:
    col1, col2 = st.columns([1, 1])
    with col1:
        st.markdown("#### Asset Class Allocation")
        ac = df.groupby("Asset_Class").agg(MV=("Market_Value", "sum")).reset_index()
        fig = px.pie(ac, values="MV", names="Asset_Class", hole=0.55,
                     color_discrete_sequence=["#2563eb", "#10b981", "#f59e0b", "#ef4444", "#8b5cf6"])
        fig.update_traces(textposition="outside", textinfo="label+percent")
        fig.update_layout(showlegend=False, height=380, margin=dict(l=10, r=10, t=10, b=10))
        st.plotly_chart(fig, width="stretch")
    with col2:
        st.markdown("#### Sub-Category Breakdown")
        sc = df.groupby("Sub_Category").agg(MV=("Market_Value", "sum")).reset_index().sort_values("MV", ascending=True)
        fig = go.Figure(go.Bar(y=sc["Sub_Category"], x=sc["MV"], orientation="h",
                                marker_color="#2563eb",
                                text=sc["MV"].apply(lambda x: fmt_inr(x)),
                                textposition="outside"))
        fig.update_layout(height=380, margin=dict(l=10, r=10, t=10, b=10),
                          xaxis=dict(title="Market Value (₹)", tickformat=",.0f"))
        st.plotly_chart(fig, width="stretch")

    st.markdown("#### Family Treemap (gain-coloured)")
    tm = df.groupby(["Holder", "AMC", "Asset_Class"]).agg(
        MV=("Market_Value", "sum"), Cost=("Cost", "sum"),
    ).reset_index()
    tm["Gain_Pct"] = (tm["MV"] / tm["Cost"] - 1) * 100
    fig = px.treemap(tm, path=[px.Constant("Family"), "Holder", "AMC", "Asset_Class"],
                     values="MV", color="Gain_Pct",
                     color_continuous_scale=[(0, "#dc2626"), (0.5, "#fef3c7"), (1, "#059669")],
                     color_continuous_midpoint=0,
                     hover_data={"Cost": ":,.0f", "MV": ":,.0f", "Gain_Pct": ":+.2f"})
    fig.update_traces(textinfo="label+value+percent parent")
    fig.update_layout(height=520, margin=dict(l=10, r=10, t=10, b=10))
    st.plotly_chart(fig, width="stretch")


# -------------------- 2. FAMILY (BY PAN) --------------------
with tabs[1]:
    pan_xirrs = aggregate_xirr(filtered_schemes, lambda s: s.pan)
    pan_summary = df.groupby("PAN").agg(
        Holder=("Holder", "first"), Schemes=("Scheme", "count"),
        Cost=("Cost", "sum"), Market_Value=("Market_Value", "sum"),
    ).reset_index()
    pan_summary["Gain"] = pan_summary["Market_Value"] - pan_summary["Cost"]
    pan_summary["Abs_Return_Pct"] = (pan_summary["Market_Value"] / pan_summary["Cost"] - 1) * 100
    pan_summary["XIRR_Pct"] = pan_summary["PAN"].map(pan_xirrs)
    pan_summary = pan_summary.sort_values("Market_Value", ascending=False)

    st.markdown("#### Per-PAN Summary")
    st.dataframe(pan_summary, width="stretch", hide_index=True,
        column_config={
            "Cost": st.column_config.NumberColumn("Cost (₹)", format="%.0f"),
            "Market_Value": st.column_config.NumberColumn("MV (₹)", format="%.0f"),
            "Gain": st.column_config.NumberColumn("Gain (₹)", format="%.0f"),
            "Abs_Return_Pct": st.column_config.NumberColumn("Abs Return", format="%.2f%%"),
            "XIRR_Pct": st.column_config.NumberColumn("XIRR", format="%.2f%%"),
        })

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("##### Market Value by PAN")
        fig = px.bar(pan_summary, x="Holder", y="Market_Value", color="Holder",
                     text=pan_summary["Market_Value"].apply(lambda x: fmt_inr(x)))
        fig.update_traces(textposition="outside")
        fig.update_layout(height=350, showlegend=False,
                          margin=dict(l=10, r=10, t=10, b=10),
                          yaxis=dict(tickformat=",.0f"))
        st.plotly_chart(fig, width="stretch")
    with col2:
        st.markdown("##### XIRR by PAN")
        chart = pan_summary.dropna(subset=["XIRR_Pct"]).copy()
        chart["Color"] = chart["XIRR_Pct"].apply(lambda x: "#059669" if x >= 0 else "#dc2626")
        fig = go.Figure(go.Bar(x=chart["Holder"], y=chart["XIRR_Pct"],
                                marker_color=chart["Color"],
                                text=chart["XIRR_Pct"].apply(lambda x: f"{x:+.2f}%"),
                                textposition="outside"))
        fig.update_layout(height=350, margin=dict(l=10, r=10, t=10, b=10),
                          yaxis=dict(title="XIRR (%)"))
        st.plotly_chart(fig, width="stretch")

    st.markdown("#### PAN × AMC Detail")
    pa_xirrs = aggregate_xirr(filtered_schemes, lambda s: (s.pan, s.amc))
    pa = df.groupby(["PAN", "AMC"]).agg(
        Holder=("Holder", "first"), Schemes=("Scheme", "count"),
        Cost=("Cost", "sum"), Market_Value=("Market_Value", "sum"),
    ).reset_index()
    pa["Gain"] = pa["Market_Value"] - pa["Cost"]
    pa["Abs_Return_Pct"] = (pa["Market_Value"] / pa["Cost"] - 1) * 100
    pa["XIRR_Pct"] = pa.apply(lambda r: pa_xirrs.get((r["PAN"], r["AMC"])), axis=1)
    st.dataframe(pa, width="stretch", hide_index=True,
        column_config={
            "Cost": st.column_config.NumberColumn("Cost (₹)", format="%.0f"),
            "Market_Value": st.column_config.NumberColumn("MV (₹)", format="%.0f"),
            "Gain": st.column_config.NumberColumn("Gain (₹)", format="%.0f"),
            "Abs_Return_Pct": st.column_config.NumberColumn("Abs Return", format="%.2f%%"),
            "XIRR_Pct": st.column_config.NumberColumn("XIRR", format="%.2f%%"),
        })


# -------------------- 3. BY AMC --------------------
with tabs[2]:
    amc_xirrs = aggregate_xirr(filtered_schemes, lambda s: s.amc)
    amc_summary = df.groupby("AMC").agg(
        Schemes=("Scheme", "count"), PANs=("PAN", "nunique"),
        Cost=("Cost", "sum"), Market_Value=("Market_Value", "sum"),
    ).reset_index()
    amc_summary["Gain"] = amc_summary["Market_Value"] - amc_summary["Cost"]
    amc_summary["Abs_Return_Pct"] = (amc_summary["Market_Value"] / amc_summary["Cost"] - 1) * 100
    amc_summary["XIRR_Pct"] = amc_summary["AMC"].map(amc_xirrs)
    amc_summary = amc_summary.sort_values("Market_Value", ascending=False)
    st.dataframe(amc_summary, width="stretch", hide_index=True,
        column_config={
            "Cost": st.column_config.NumberColumn("Cost (₹)", format="%.0f"),
            "Market_Value": st.column_config.NumberColumn("MV (₹)", format="%.0f"),
            "Gain": st.column_config.NumberColumn("Gain (₹)", format="%.0f"),
            "Abs_Return_Pct": st.column_config.NumberColumn("Abs Return", format="%.2f%%"),
            "XIRR_Pct": st.column_config.NumberColumn("XIRR", format="%.2f%%"),
        })

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("##### Allocation by AMC")
        fig = px.pie(amc_summary, values="Market_Value", names="AMC", hole=0.4)
        fig.update_traces(textposition="inside", textinfo="percent+label")
        fig.update_layout(height=400, margin=dict(l=10, r=10, t=10, b=10), showlegend=False)
        st.plotly_chart(fig, width="stretch")
    with col2:
        st.markdown("##### XIRR by AMC")
        chart = amc_summary.dropna(subset=["XIRR_Pct"]).sort_values("XIRR_Pct")
        fig = go.Figure(go.Bar(y=chart["AMC"], x=chart["XIRR_Pct"], orientation="h",
                                marker_color=chart["XIRR_Pct"].apply(
                                    lambda x: "#059669" if x >= 0 else "#dc2626"),
                                text=chart["XIRR_Pct"].apply(lambda x: f"{x:+.2f}%"),
                                textposition="outside"))
        fig.update_layout(height=400, margin=dict(l=10, r=10, t=10, b=10),
                          xaxis=dict(title="XIRR (%)"))
        st.plotly_chart(fig, width="stretch")


# -------------------- 4. SCHEMES --------------------
with tabs[3]:
    st.markdown(f"#### {len(df)} schemes (sortable)")
    show_df = df.sort_values("Market_Value", ascending=False)
    st.dataframe(
        show_df[["PAN", "Holder", "AMC", "Scheme", "Asset_Class", "Sub_Category",
                 "Cost", "Market_Value", "Gain", "Abs_Return_Pct", "XIRR_Pct",
                 "First_Investment", "SIP_Status", "Folio"]],
        width="stretch", hide_index=True, height=480,
        column_config={
            "Cost": st.column_config.NumberColumn("Cost (₹)", format="%.0f"),
            "Market_Value": st.column_config.NumberColumn("MV (₹)", format="%.0f"),
            "Gain": st.column_config.NumberColumn("Gain (₹)", format="%.0f"),
            "Abs_Return_Pct": st.column_config.NumberColumn("Abs Return", format="%.2f%%"),
            "XIRR_Pct": st.column_config.NumberColumn("XIRR", format="%.2f%%"),
            "Scheme": st.column_config.TextColumn("Scheme", width="large"),
            "First_Investment": st.column_config.DateColumn("Started"),
        },
    )


# -------------------- 5. CAPITAL GAINS --------------------
with tabs[4]:
    st.caption("FIFO-matched realized capital gains. Tax computed using FY 2024-25 onwards regime "
               "(post-July 2024). Switches and lateral shifts are treated as taxable redemptions.")

    # Compute realized + unrealized
    all_realized: list = []
    all_unrealized: list = []
    for s in filtered_schemes:
        r, u = compute_capital_gains(s, st.session_state.tax_rates)
        all_realized.extend(r)
        all_unrealized.extend(u)

    if not all_realized and not all_unrealized:
        st.info("No capital-gains data yet — possibly all schemes are too new.")
    else:
        # Headline numbers
        realized_gain = sum(l.gain for l in all_realized)
        unrealized_gain = sum(l.gain for l in all_unrealized)
        ltcg_unreal = sum(l.gain for l in all_unrealized if l.is_long_term)
        stcg_unreal = sum(l.gain for l in all_unrealized if not l.is_long_term)

        cg1, cg2, cg3, cg4 = st.columns(4)
        cg1.metric("Realized Gain (lifetime)", fmt_inr(realized_gain),
                   delta=f"{len(all_realized)} lots", delta_color="off")
        cg2.metric("Unrealized Gain", fmt_inr(unrealized_gain),
                   delta=f"{len(all_unrealized)} lots", delta_color="off")
        cg3.metric("Unrealized LTCG", fmt_inr(ltcg_unreal))
        cg4.metric("Unrealized STCG", fmt_inr(stcg_unreal))

        # Realized by FY
        if all_realized:
            tax_summary = compute_tax_for_lots(
                all_realized, st.session_state.tax_rates,
                st.session_state.debt_slab / 100,
            )
            st.markdown("#### Realized Gains by Financial Year")
            tax_rows = pd.DataFrame([{
                "FY": fy, "Total Gain": s.total_gain,
                "Equity STCG": s.equity_stcg, "Equity LTCG": s.equity_ltcg_gross,
                "LTCG Taxable": s.equity_ltcg_taxable,
                "Debt STCG": s.debt_stcg, "Debt LTCG": s.debt_ltcg,
                "Total Tax": s.total_tax,
            } for fy, s in sorted(tax_summary.items())])
            st.dataframe(tax_rows, width="stretch", hide_index=True,
                         column_config={c: st.column_config.NumberColumn(c, format="%.0f")
                                        for c in tax_rows.columns if c != "FY"})

            # Lot-level table
            with st.expander("📑 Lot-level realized gains (ITR-format export)", expanded=False):
                lots_df = pd.DataFrame([{
                    "FY": l.fy, "Holder": l.holder, "Scheme": l.scheme,
                    "Buy Date": l.buy_date, "Sell Date": l.sell_date,
                    "Units": round(l.units, 4),
                    "Buy Price": round(l.buy_price, 4), "Sell Price": round(l.sell_price, 4),
                    "Cost": round(l.cost, 2), "Proceeds": round(l.proceeds, 2),
                    "Gain": round(l.gain, 2), "Days": l.holding_days,
                    "Term": "LTCG" if l.is_long_term else "STCG",
                    "Asset": l.tax_category.title(),
                } for l in all_realized])
                st.dataframe(lots_df.sort_values(["FY", "Sell Date"], ascending=[False, False]),
                             width="stretch", hide_index=True, height=400)
                csv = lots_df.to_csv(index=False).encode("utf-8")
                st.download_button("⬇️ Download lot-level CSV", csv,
                                   "realized_gains.csv", "text/csv")

        # Unrealized: bar chart by scheme
        if all_unrealized:
            st.markdown("#### Unrealized Position — by Scheme")
            ur_by_scheme = defaultdict(lambda: {"cost": 0, "value": 0, "gain": 0,
                                                 "ltcg": 0, "stcg": 0, "holder": "", "amc": ""})
            for l in all_unrealized:
                d = ur_by_scheme[l.scheme]
                d["cost"] += l.cost; d["value"] += l.current_value; d["gain"] += l.gain
                if l.is_long_term: d["ltcg"] += l.gain
                else: d["stcg"] += l.gain
                d["holder"] = l.holder; d["amc"] = l.amc
            ur_df = pd.DataFrame([{
                "Scheme": k[:55] + ("…" if len(k) > 55 else ""),
                "Holder": v["holder"], "AMC": v["amc"],
                "Cost": v["cost"], "Current Value": v["value"],
                "Gain": v["gain"], "LTCG": v["ltcg"], "STCG": v["stcg"],
            } for k, v in ur_by_scheme.items()]).sort_values("Current Value", ascending=False)
            st.dataframe(ur_df, width="stretch", hide_index=True, height=350,
                         column_config={c: st.column_config.NumberColumn(c, format="%.0f")
                                        for c in ["Cost", "Current Value", "Gain", "LTCG", "STCG"]})


# -------------------- 6. TAX PREVIEW --------------------
with tabs[5]:
    st.caption("Simulate tax if you redeem holdings today. Adjust the rates below and the debt slab.")

    col1, col2, col3 = st.columns([1, 1, 1])
    with col1:
        eq_ltcg_rate = st.number_input("Equity LTCG rate %", 0.0, 30.0,
                                        st.session_state.tax_rates.equity_ltcg_rate * 100, 0.5)
        eq_stcg_rate = st.number_input("Equity STCG rate %", 0.0, 50.0,
                                        st.session_state.tax_rates.equity_stcg_rate * 100, 0.5)
    with col2:
        debt_ltcg_rate = st.number_input("Debt LTCG rate %", 0.0, 30.0,
                                          st.session_state.tax_rates.debt_ltcg_rate * 100, 0.5)
        debt_slab = st.number_input("Debt STCG slab %", 0.0, 50.0,
                                     st.session_state.debt_slab, 0.5)
    with col3:
        ltcg_exempt = st.number_input("Equity LTCG exemption (₹)", 0, 500_000,
                                       int(st.session_state.tax_rates.equity_ltcg_exemption), 25_000)

    rates = TaxRates(
        equity_ltcg_rate=eq_ltcg_rate / 100, equity_ltcg_exemption=ltcg_exempt,
        equity_stcg_rate=eq_stcg_rate / 100, debt_ltcg_rate=debt_ltcg_rate / 100,
        debt_stcg_slab=debt_slab / 100,
    )

    # Persist
    st.session_state.tax_rates = rates
    st.session_state.debt_slab = debt_slab

    proceeds, total_tax, net, summary = simulate_full_redemption_tax(
        filtered_schemes, rates, debt_slab / 100)

    st.markdown("### Hypothetical Full Redemption Today")
    t1, t2, t3, t4 = st.columns(4)
    t1.metric("Total Proceeds", fmt_inr(proceeds))
    t2.metric("Total Tax", fmt_inr(total_tax),
              delta=f"-{total_tax/proceeds*100:.2f}%" if proceeds else "—",
              delta_color="inverse")
    t3.metric("Total Gain", fmt_inr(summary.total_gain))
    t4.metric("Net to You", fmt_inr(net))

    st.markdown("#### Tax Breakdown")
    breakdown = pd.DataFrame([
        {"Component": "Equity STCG", "Gain": summary.equity_stcg,
         "Taxable": summary.equity_stcg, "Rate %": eq_stcg_rate, "Tax": summary.equity_stcg_tax},
        {"Component": "Equity LTCG", "Gain": summary.equity_ltcg_gross,
         "Taxable": summary.equity_ltcg_taxable, "Rate %": eq_ltcg_rate, "Tax": summary.equity_ltcg_tax},
        {"Component": "Debt STCG", "Gain": summary.debt_stcg,
         "Taxable": summary.debt_stcg, "Rate %": debt_slab, "Tax": summary.debt_stcg_tax},
        {"Component": "Debt LTCG", "Gain": summary.debt_ltcg,
         "Taxable": summary.debt_ltcg, "Rate %": debt_ltcg_rate, "Tax": summary.debt_ltcg_tax},
    ])
    st.dataframe(breakdown, width="stretch", hide_index=True,
                 column_config={c: st.column_config.NumberColumn(c, format="%.0f")
                                for c in ["Gain", "Taxable", "Tax"]})

    # Insight: harvest LTCG up to exemption
    unreal_ltcg = sum(
        l.gain for s in filtered_schemes
        for l in compute_capital_gains(s, rates)[1]
        if l.is_long_term and l.tax_category == "equity"
    )
    if unreal_ltcg > 0 and ltcg_exempt > 0:
        st.markdown(f"<div class='insight'>💡 <b>Tax-harvesting tip:</b> You have ₹{unreal_ltcg:,.0f} "
                    f"of unrealized equity LTCG. You can realize up to ₹{ltcg_exempt:,.0f} "
                    f"per FY tax-free (exemption limit). Selling and immediately re-buying lets you "
                    f"reset the cost basis without paying tax.</div>",
                    unsafe_allow_html=True)


# -------------------- 7. ALLOCATION DRIFT --------------------
with tabs[6]:
    st.caption("Set your target allocation by asset class. The dashboard shows drift and the "
               "rupee amount needed to rebalance.")

    classes_seen = sorted(set([categorize_scheme(s.scheme_name).asset_class for s in filtered_schemes]))
    default_targets = st.session_state.allocation_targets

    cols = st.columns(min(5, max(1, len(classes_seen))))
    new_targets = {}
    for i, c in enumerate(classes_seen):
        with cols[i % len(cols)]:
            new_targets[c] = st.number_input(
                f"{c} target %", 0.0, 100.0,
                float(default_targets.get(c, 0.0)), 5.0,
                key=f"tgt_{c}",
            )
    total_target = sum(new_targets.values())
    if abs(total_target - 100) > 0.5:
        st.markdown(f"<div class='warning'>⚠️ Targets sum to {total_target:.1f}%. Adjust to 100%.</div>",
                    unsafe_allow_html=True)
    st.session_state.allocation_targets = new_targets

    drift = compute_drift(filtered_schemes, new_targets)
    drift_df = pd.DataFrame([{
        "Asset Class": d.asset_class, "Current Value": d.current_value,
        "Current %": d.current_pct, "Target %": d.target_pct,
        "Drift %": d.drift_pct, "Action": "Buy" if d.rebalance_amount > 0 else "Sell",
        "Rebalance ₹": abs(d.rebalance_amount),
    } for d in drift])

    st.markdown("#### Drift vs Target")
    st.dataframe(drift_df, width="stretch", hide_index=True,
                 column_config={
                     "Current Value": st.column_config.NumberColumn(format="%.0f"),
                     "Current %": st.column_config.NumberColumn(format="%.2f%%"),
                     "Target %": st.column_config.NumberColumn(format="%.2f%%"),
                     "Drift %": st.column_config.NumberColumn(format="%.2f%%"),
                     "Rebalance ₹": st.column_config.NumberColumn(format="%.0f"),
                 })

    # Drift chart
    fig = go.Figure()
    fig.add_trace(go.Bar(name="Current %", x=drift_df["Asset Class"], y=drift_df["Current %"],
                          marker_color="#2563eb"))
    fig.add_trace(go.Bar(name="Target %", x=drift_df["Asset Class"], y=drift_df["Target %"],
                          marker_color="#94a3b8"))
    fig.update_layout(barmode="group", height=380,
                      margin=dict(l=10, r=10, t=10, b=10),
                      yaxis=dict(title="%"),
                      legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
    st.plotly_chart(fig, width="stretch")

    # Rebalance suggestions
    big_drifts = [d for d in drift if abs(d.drift_pct) > 2]
    if big_drifts:
        st.markdown("#### Rebalance Suggestions")
        for d in big_drifts:
            arrow = "📈 Buy" if d.rebalance_amount > 0 else "📉 Sell"
            sign = "overweight" if d.drift_pct > 0 else "underweight"
            st.markdown(
                f"<div class='insight'>{arrow} <b>{fmt_inr(abs(d.rebalance_amount))}</b> "
                f"of {d.asset_class} — currently {d.current_pct:.1f}% "
                f"({sign} by {abs(d.drift_pct):.1f}%, target {d.target_pct:.1f}%)</div>",
                unsafe_allow_html=True)


# -------------------- 8. SIP HEALTH --------------------
with tabs[7]:
    st.caption("Active and cancelled SIPs across all your folios. "
               "An SIP is flagged 'cancelled' if the CAS recorded a cancellation event.")

    sips = analyze_sips(filtered_schemes)
    if not sips:
        st.info("No SIP transactions detected in the selected schemes.")
    else:
        sip_df = pd.DataFrame([{
            "PAN": s.pan, "Holder": s.holder, "AMC": s.amc, "Scheme": s.scheme,
            "Status": s.status.title(),
            "Typical SIP ₹": s.typical_amount, "SIP Count": s.sip_count,
            "Last SIP": s.last_sip_date,
            "Days Since Last": s.days_since_last,
            "Cancelled On": s.cancel_date,
        } for s in sips])

        # Counts
        active = sip_df[sip_df["Status"] == "Active"]
        cancelled = sip_df[sip_df["Status"] == "Cancelled"]

        s1, s2, s3, s4 = st.columns(4)
        s1.metric("Active SIPs", len(active))
        s2.metric("Cancelled SIPs", len(cancelled))
        s3.metric("Monthly Outflow (active)", fmt_inr(active["Typical SIP ₹"].sum()))
        s4.metric("Total SIP Instalments", int(sip_df["SIP Count"].sum()))

        st.markdown("#### Active SIPs")
        st.dataframe(active.sort_values("Typical SIP ₹", ascending=False),
                     width="stretch", hide_index=True,
                     column_config={
                         "Typical SIP ₹": st.column_config.NumberColumn(format="%.0f"),
                         "Last SIP": st.column_config.DateColumn(),
                     })

        if len(cancelled):
            st.markdown("#### Cancelled SIPs")
            st.dataframe(cancelled, width="stretch", hide_index=True,
                         column_config={
                             "Typical SIP ₹": st.column_config.NumberColumn(format="%.0f"),
                             "Last SIP": st.column_config.DateColumn(),
                             "Cancelled On": st.column_config.DateColumn(),
                         })

        # Stale SIPs warning (>40 days since last)
        stale = active[active["Days Since Last"].fillna(0) > 40]
        if len(stale):
            st.markdown(f"<div class='warning'>⚠️ <b>{len(stale)} 'active' SIPs haven't run in 40+ days</b> "
                        f"— possibly bounced, on hold, or recently registered with no instalment yet.</div>",
                        unsafe_allow_html=True)


# -------------------- 9. CASH FLOWS --------------------
with tabs[8]:
    cf_rows = []
    for s in filtered_schemes:
        for t in s.cashflows:
            cf_rows.append({
                "Date": t.date, "Amount": t.signed_amount, "Kind": t.kind,
                "PAN": s.pan, "Holder": holders.get(s.pan, ""),
                "AMC": s.amc, "Scheme": s.scheme_name,
            })
    cf_df = pd.DataFrame(cf_rows)
    if cf_df.empty:
        st.info("No transactions in selection.")
    else:
        cf_df["Date"] = pd.to_datetime(cf_df["Date"])
        cf_df["Month"] = cf_df["Date"].dt.to_period("M").dt.to_timestamp()
        cf_df["Outflow"] = cf_df["Amount"].apply(lambda x: -x if x < 0 else 0)
        cf_df["Inflow"] = cf_df["Amount"].apply(lambda x: x if x > 0 else 0)

        # Yearly summary
        cf_df["Year"] = cf_df["Date"].dt.year
        yearly = cf_df.groupby("Year").agg(
            Invested=("Outflow", "sum"), Redeemed=("Inflow", "sum"),
            Transactions=("Date", "count"),
        ).reset_index()
        yearly["Net"] = yearly["Invested"] - yearly["Redeemed"]

        col1, col2 = st.columns([2, 1])
        with col1:
            st.markdown("#### Monthly Cash Flow")
            monthly = cf_df.groupby("Month").agg(
                Invested=("Outflow", "sum"), Redeemed=("Inflow", "sum"),
            ).reset_index()
            monthly["Cum_Net"] = (monthly["Invested"] - monthly["Redeemed"]).cumsum()

            fig = go.Figure()
            fig.add_trace(go.Bar(x=monthly["Month"], y=monthly["Invested"],
                                  name="Invested", marker_color="#2563eb", opacity=0.7))
            fig.add_trace(go.Bar(x=monthly["Month"], y=-monthly["Redeemed"],
                                  name="Redeemed", marker_color="#ef4444", opacity=0.7))
            fig.add_trace(go.Scatter(x=monthly["Month"], y=monthly["Cum_Net"],
                                      name="Cumulative Net Invested",
                                      line=dict(color="#1e40af", width=3), yaxis="y2"))
            fig.update_layout(height=420, hovermode="x unified", barmode="relative",
                              margin=dict(l=10, r=10, t=10, b=10),
                              yaxis=dict(title="Monthly (₹)", tickformat=",.0f"),
                              yaxis2=dict(title="Cumulative (₹)", overlaying="y",
                                          side="right", tickformat=",.0f"),
                              legend=dict(orientation="h", yanchor="bottom",
                                          y=1.02, xanchor="right", x=1))
            st.plotly_chart(fig, width="stretch")

        with col2:
            st.markdown("#### Yearly")
            st.dataframe(yearly, width="stretch", hide_index=True,
                         column_config={
                             "Invested": st.column_config.NumberColumn(format="%.0f"),
                             "Redeemed": st.column_config.NumberColumn(format="%.0f"),
                             "Net": st.column_config.NumberColumn(format="%.0f"),
                         })


# -------------------- 10. EXPORT --------------------
with tabs[9]:
    st.markdown("### Export Your Portfolio")
    st.caption("Take your data with you — save to local database, export Excel, or JSON snapshot.")

    e1, e2, e3 = st.columns(3)
    with e1:
        st.markdown("#### 📊 Excel Workbook")
        st.markdown("Multi-sheet workbook: per-PAN, per-AMC, per-scheme, asset allocation, "
                    "**realized gains by FY (ITR-format), unrealized gains, tax summary**.")
        if st.button("Generate Excel", type="primary", width="stretch"):
            with st.spinner("Building workbook..."):
                xlsx_bytes = export_excel()
            st.download_button(
                "⬇️ Download portfolio.xlsx",
                xlsx_bytes,
                file_name=f"portfolio_{date.today().isoformat()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
            )
    with e2:
        st.markdown("#### 💾 JSON State Snapshot")
        st.markdown("Full parsed state including transactions and your settings (targets, tax rates). "
                    "Re-import without re-parsing PDFs.")
        if st.button("Generate JSON", type="primary", width="stretch"):
            with st.spinner("Serializing..."):
                json_bytes = export_state_json()
            st.download_button(
                "⬇️ Download state.json",
                json_bytes,
                file_name=f"portfolio_state_{date.today().isoformat()}.json",
                mime="application/json",
                width="stretch",
            )
    with e3:
        st.markdown("#### 💿 Save to Database")
        st.markdown("Persist portfolio to SQLite/PostgreSQL. Re-load anytime without re-parsing PDFs.")
        portfolio_name = st.text_input("Portfolio name", value=f"Portfolio {date.today().isoformat()}",
                                     key="db_portfolio_name", label_visibility="collapsed")
        if st.button("Save to DB", type="primary", width="stretch",
                    disabled=not uploaded):
            try:
                session = get_db()
                tax_rates_dict = {
                    "equity_ltcg_rate": st.session_state.tax_rates.equity_ltcg_rate,
                    "equity_ltcg_exemption": st.session_state.tax_rates.equity_ltcg_exemption,
                    "equity_stcg_rate": st.session_state.tax_rates.equity_stcg_rate,
                    "debt_ltcg_rate": st.session_state.tax_rates.debt_ltcg_rate,
                    "debt_stcg_slab": st.session_state.debt_slab / 100,
                    "equity_holding_threshold_days": st.session_state.tax_rates.equity_holding_threshold_days,
                    "debt_holding_threshold_days": st.session_state.tax_rates.debt_holding_threshold_days,
                }
                save_portfolio(
                    session, portfolio_name,
                    st.session_state.parsed_files, tax_rates_dict,
                    st.session_state.allocation_targets, st.session_state.scheme_overrides,
                    st.session_state.pan_holders,
                )
                st.success("Saved to database!")
                st.rerun()
            except Exception as e:
                st.error(f"Save failed: {e}")


# Footer
st.divider()
st.caption(
    "💡 XIRR uses real cash-flow dates with switches/lateral shifts excluded "
    "(intra-portfolio movements). Capital gains are FIFO-matched. "
    "Tax rates default to FY 2024-25 onwards regime and are user-adjustable. "
    "All processing happens in your browser session — no data persisted server-side."
)
